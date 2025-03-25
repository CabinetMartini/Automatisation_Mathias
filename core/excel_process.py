import openpyxl
import datetime
import locale


def split_date(date):
    month, year = date.split(" ")
    return month, year


def change_cell(file_xlsx, donnes):
    print(donnes)

    data_insert = {
        "Chiffre d'affaires net": float(donnes["TOTAL PRODUITS NET"]["TAC"]),
        "Transactions sur place": float(donnes["Sur place"]["TAC"]) - float(donnes["Kiosk sur place"]["TAC"]),
        "Transactions à emporter": float(donnes["A emporter"]["TAC"]) - float(donnes["Kiosk à emporter"]["TAC"]),
        "Transactions Mc Drive": float(donnes["McDrive"]["TAC"]),
        "Transactions kiosque": float(donnes["TOTAL Kiosk"]["TAC"]),
        "Transactions totales": float(donnes["TOTAL"]["TAC"]),
        "Paniers moyens sur place": float(donnes["Sur place"]["P.M."]),
        "Paniers moyens à emporter": float(donnes["A emporter"]["P.M."]),
        "Paniers moyens total": float(donnes["TOTAL"]["P.M."]),
    }
    print(data_insert)

    month, year =split_date(donnes["Date"])
    tab_name = donnes["Nom du fichier"].replace(".pdf", "")

    wb = openpyxl.load_workbook(file_xlsx)
    ws = wb.active

    rows = find_range_val_data_insert(ws)
    col = find_name_col(ws,tab_name)

    # Pour chaque clé du dictionnaire, on récupère les coordonnées de début et de fin de la plage fusionnée
    for key, value in rows.items():
        # Ici, value est une chaîne du type "A11:A23"
        coord1, coord2 = value.split(":")
        # On extrait la partie numérique en ignorant le premier caractère (la lettre)
        min_number = coord1[1:]
        max_number = coord2[1:]
        # On reconstruit la référence en concaténant la nouvelle colonne et le numéro
        min_cell = col + min_number
        max_cell = col + max_number
        rows[key] = (min_cell, max_cell)
        print(f"Pour '{key}': min_cell: {min_cell}, max_cell: {max_cell}")

    # On cherche les données des plages fusionnées
    for key, value in rows.items():
        min_cell, max_cell = value
        #print(f"Pour '{key}': min_cell: {min_cell}, max_cell: {max_cell}")
        for i in range(int(min_cell[1:]), int(max_cell[1:]) + 1):
            full_month, month_number_str = normalize_month(month)
            if (ws[f"{min_cell[0]}{i}"].value == month or ws[f"{min_cell[0]}{i}"].value == month.upper() or ws[f"{min_cell[0]}{i}"].value == full_month or ws[f"{min_cell[0]}{i}"].value == month_number_str):                
                current_col_letter = min_cell[0]
                next_col = chr(ord(current_col_letter) + 1)

                if (f"{next_col}{min_cell}") == year:
                    target_cell = f"{next_col}{i}"
                    ws[target_cell].value = data_insert[key]
                    print(f"Valeur {data_insert[key]} placée en {target_cell}") 
                elif (f"{chr(ord(current_col_letter) + 2)}{min_cell}") == year:
                    next_col = chr(ord(current_col_letter) + 1)
                    target_cell = f"{next_col}{i}"
                    ws[target_cell].value = data_insert[key]
                    print(f"Valeur {data_insert[key]} placée en {target_cell}")

            #print(f"Valeur actuelle : {ws[f'{min_cell[0]}{i}'].value}")
        
    wb.save(file_xlsx)

    wb.close()
    return data_insert



def find_range_val_data_insert(ws):
    range_lines ={
        "Chiffre d'affaires net": "",
        "Transactions sur place": "",
        "Transactions à emporter": "",
        "Transactions Mc Drive": "",
        "Transactions kiosque": "",
        "Transactions totales": "",
        "Paniers moyens sur place": "",
        "Paniers moyens à emporter": "",
        "Paniers moyens total": "",
    }

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        cell_value = ws.cell(row=min_row, column=min_col).value

        for key in range_lines.keys():
            if key == cell_value:
                range_lines[key] = (merged_range.coord)
                #range_lines[key].append(min_row)
                #range_lines[key].append(max_row)
                break
        #print(f"Plage fusionnée {merged_range.coord} : de la ligne {min_row} à la ligne {max_row}, valeur associée : {cell_value}")
    
    print(range_lines)
    return range_lines


def find_name_col(ws, name_col):
    print(name_col)
    for row in ws.iter_rows(min_row=9, max_row=9):
        for cell in row:
            if cell.value == name_col:
                return cell.column_letter



def normalize_month(month_value):
    """
    Normalise la valeur du mois qui peut être soit un nom complet (ex. "Février")
    soit un numéro sous forme de chaîne (ex. "02"). Renvoie un tuple composé du nom
    complet du mois (en français) et du numéro du mois formaté sur 2 chiffres.
    """
    # Dictionnaire de correspondance : clé en minuscule
    month_mapping = {
        "1": 1, "01": 1, "janvier": 1,
        "2": 2, "02": 2, "février": 2, "fevrier": 2,
        "3": 3, "03": 3, "mars": 3,
        "4": 4, "04": 4, "avril": 4,
        "5": 5, "05": 5, "mai": 5,
        "6": 6, "06": 6, "juin": 6,
        "7": 7, "07": 7, "juillet": 7,
        "8": 8, "08": 8, "août": 8, "aout": 8,
        "9": 9, "09": 9, "septembre": 9,
        "10": 10, "octobre": 10,
        "11": 11, "novembre": 11,
        "12": 12, "décembre": 12, "decembre": 12
    }
    
    # Normalisation de la chaîne
    month_norm = str(month_value).strip().lower()
    
    if month_norm in month_mapping:
        month_number = month_mapping[month_norm]
    else:
        # On tente une conversion directe en entier
        try:
            month_number = int(month_norm)
        except ValueError:
            raise ValueError(f"Le mois '{month_value}' n'est pas reconnu.")
    
    # Optionnel : tenter de définir la locale en français pour obtenir le nom du mois en français
    try:
        locale.setlocale(locale.LC_TIME, "fr_FR.UTF-8")
    except locale.Error:
        pass  # Si la locale française n'est pas disponible, le nom risque d'être en anglais
    
    # Obtenir le nom complet du mois (ex: "février") et le numéro au format 2 chiffres
    full_month = datetime.date(1900, month_number, 1).strftime('%B')
    month_str = f"{month_number:02d}"
    return full_month, month_str

#print(f"full_month: {full_month}, month_number_str: {month_number_str}")

