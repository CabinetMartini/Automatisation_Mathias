import pandas as pd
import openpyxl

def split_date(date):
    month, year = date.split(" ")
    return month, year


def change_cell(file_xlsx, donnes):

    data_insert = {
        "Chiffre d'affaires net": donnes["TOTAL PRODUITS NET"]["TAC"],
        "Transactions sur place": donnes["Sur place"]["TAC"],
        "Transactions à emporter": donnes["A emporter"]["TAC"],
        "Transactions Mc Drive": donnes["McDrive"]["TAC"],
        "Transactions kiosque": donnes["TOTAL Kiosk"]["TAC"],
        "Transactions totales": donnes["TOTAL"]["TAC"],
        "Paniers moyens sur place": donnes["Sur place"]["P.M."],
        "Paniers moyens à emporter": donnes["A emporter"]["P.M."],
        "Paniers moyens total": donnes["TOTAL"]["P.M."],
    }

    month, year =split_date(donnes["Date"])
    #print(f"month: {month}, year: {year}")
    tab_name = donnes["Nom du fichier"].replace(" FC2.pdf", "")
    #print(f"tab_name: {tab_name}")

    wb = openpyxl.load_workbook(file_xlsx)
    ws = wb.active


    rows = find_range_val_data_insert(ws)
    print(f"rows: {rows}")
    col = find_name_col(ws,tab_name)
    print(f"col: {col}")

    # Pour chaque clé du dictionnaire, on récupère les coordonnées de début et de fin de la plage fusionnée
    for key, value in rows.items():
        # Ici, value est une chaîne du type "A11:A23"
        coord1, coord2 = value.split(":")
        # On extrait la partie numérique en ignorant le premier caractère (la lettre)
        min_number = coord1[1:]
        max_number = coord2[1:]
        # On reconstruit la référence en concaténant la nouvelle colonne et le numéro
        min_cell = col + min_number
        max_cell = col + max_number
        rows[key] = (min_cell, max_cell)
        print(f"Pour '{key}': min_cell: {min_cell}, max_cell: {max_cell}")

    # On cherche les données des plages fusionnées
    for key, value in rows.items():
        min_cell, max_cell = value
        print(f"Pour '{key}': min_cell: {min_cell}, max_cell: {max_cell}")
        for i in range(int(min_cell[1:]), int(max_cell[1:]) + 1):
            #print(f"min : {int(min_cell[1:])}")
            #print(f"max : {int(max_cell[1:])}")
            print(f"month: {month}, {month.upper()}")
            if ws[f"{min_cell[0]}{i}"].value == month or ws[f"{min_cell[0]}{i}"].value == month.upper(): 
                print(f"Trouvé à la ligne {i} pour {key}")
                current_col_letter = min_cell[0]
                print(f"current_col_letter: {current_col_letter}")
                next_col = chr(ord(current_col_letter) + 1)
                print(f"next_col: {next_col}")
                target_cell = f"{next_col}{i}"
                ws[target_cell].value = data_insert[key]
                print(f"Valeur {data_insert[key]} placée en {target_cell}")


            print(f"Valeur actuelle : {ws[f'{min_cell[0]}{i}'].value}")
        
    wb.save(file_xlsx)

    wb.close()
    return data_insert



def find_range_val_data_insert(ws):
    range_lines ={
        "Chiffre d'affaires net": "",
        "Transactions sur place": "",
        "Transactions à emporter": "",
        "Transactions Mc Drive": "",
        "Transactions kiosque": "",
        "Transactions totales": "",
        "Paniers moyens sur place": "",
        "Paniers moyens à emporter": "",
        "Paniers moyens total": "",
    }

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        cell_value = ws.cell(row=min_row, column=min_col).value

        for key in range_lines.keys():
            if key == cell_value:
                range_lines[key] = (merged_range.coord)
                #range_lines[key].append(min_row)
                #range_lines[key].append(max_row)
                break
        #print(f"Plage fusionnée {merged_range.coord} : de la ligne {min_row} à la ligne {max_row}, valeur associée : {cell_value}")
    
    print(range_lines)
    return range_lines


def find_name_col(ws, name_col):
    for row in ws.iter_rows(min_row=9, max_row=9):
        for cell in row:
            if cell.value == name_col:
                return cell.column_letter






if __name__ == "__main__":
    path = "File.xlsx"
    data = change_cell(path)
    print(data)

